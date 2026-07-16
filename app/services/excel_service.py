"""Generación del XLSX del reporte de facturas emitidas (openpyxl)."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.afip.constants import TIPOS_COMPROBANTE
from app.models.cliente import Cliente
from app.models.comprobante import Comprobante
from app.models.empresa import Empresa
from app.services.estado_cuenta import EstadoCuenta
from app.services.reportes import calcular_totales

RELLENO_ENCABEZADO = PatternFill(
    start_color="343A40", end_color="343A40", fill_type="solid"
)
FUENTE_ENCABEZADO = Font(bold=True, color="FFFFFF")
FUENTE_TITULO = Font(bold=True, size=14)
FUENTE_TOTALES = Font(bold=True)


def generar_excel_reporte_facturas(
    comprobantes: list[Comprobante],
    *,
    fecha_inicio: str,
    fecha_fin: str,
    cliente: Cliente | None = None,
    empresa: Empresa | None = None,
) -> bytes:
    """Genera el XLSX del reporte de facturas emitidas y devuelve los bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de facturas"

    mostrar_empresa = empresa is None
    encabezado = ["Fecha", "Tipo", "Comprobante", "Cliente"]
    if mostrar_empresa:
        encabezado.append("Empresa")
    encabezado += ["Neto", "IVA", "Total", "CAE", "Vto. CAE"]
    n_cols = len(encabezado)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value="Reporte de facturas emitidas").font = FUENTE_TITULO

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(row=2, column=1, value=f"Período: {fecha_inicio} al {fecha_fin}")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    ws.cell(
        row=3,
        column=1,
        value=(
            f"Cliente: {cliente.razon_social if cliente else 'Todos los clientes'}  |  "
            f"Empresa: {empresa.razon_social if empresa else 'Todas las empresas'}"
        ),
    )

    fila_encabezado = 5
    for col, titulo in enumerate(encabezado, start=1):
        celda = ws.cell(row=fila_encabezado, column=col, value=titulo)
        celda.font = FUENTE_ENCABEZADO
        celda.fill = RELLENO_ENCABEZADO

    if not comprobantes:
        ws.merge_cells(
            start_row=fila_encabezado + 1,
            start_column=1,
            end_row=fila_encabezado + 1,
            end_column=n_cols,
        )
        ws.cell(
            row=fila_encabezado + 1,
            column=1,
            value="No se encontraron comprobantes para los filtros seleccionados.",
        )
    else:
        fila = fila_encabezado + 1
        for c in comprobantes:
            col = 1
            celda_fecha = ws.cell(
                row=fila,
                column=col,
                value=datetime.strptime(c.fecha, "%Y-%m-%d").date(),
            )
            celda_fecha.number_format = "dd/mm/yyyy"
            col += 1
            ws.cell(
                row=fila,
                column=col,
                value=TIPOS_COMPROBANTE.get(c.tipo_cbte, str(c.tipo_cbte)),
            )
            col += 1
            ws.cell(row=fila, column=col, value=f"{c.punto_venta:04d}-{c.numero:08d}")
            col += 1
            ws.cell(row=fila, column=col, value=c.cliente.razon_social)
            col += 1
            if mostrar_empresa:
                ws.cell(row=fila, column=col, value=c.empresa.razon_social)
                col += 1
            for valor in (c.importe_neto, c.importe_iva, c.importe_total):
                celda = ws.cell(row=fila, column=col, value=float(valor))
                celda.number_format = "#,##0.00"
                col += 1
            ws.cell(row=fila, column=col, value=c.cae or "")
            col += 1
            ws.cell(row=fila, column=col, value=c.cae_vencimiento or "")
            fila += 1

        totales = calcular_totales(comprobantes)
        col_totales_label = 4 + (1 if mostrar_empresa else 0)
        ws.cell(row=fila, column=col_totales_label, value="Totales:").font = (
            FUENTE_TOTALES
        )
        col = col_totales_label + 1
        for clave in ("neto", "iva", "total"):
            celda = ws.cell(row=fila, column=col, value=totales[clave])
            celda.number_format = "#,##0.00"
            celda.font = FUENTE_TOTALES
            col += 1

    anchos = {1: 12, 2: 16, 3: 16, 4: 30}
    if mostrar_empresa:
        anchos[5] = 30
    for col in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = anchos.get(col, 16)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generar_excel_estado_cuenta(estado: EstadoCuenta) -> bytes:
    """Genera el XLSX del estado de cuenta por cliente y devuelve los bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de cuenta"

    encabezado = ["Fecha", "Tipo", "Comprobante", "Importe", "Saldo"]
    n_cols = len(encabezado)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value="Estado de cuenta").font = FUENTE_TITULO

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(
        row=2, column=1, value=f"Período: {estado.fecha_inicio} al {estado.fecha_fin}"
    )

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    ws.cell(row=3, column=1, value=f"Cliente: {estado.cliente.razon_social}")

    fila_encabezado = 5
    for col, titulo in enumerate(encabezado, start=1):
        celda = ws.cell(row=fila_encabezado, column=col, value=titulo)
        celda.font = FUENTE_ENCABEZADO
        celda.fill = RELLENO_ENCABEZADO

    fila = fila_encabezado + 1
    ws.cell(row=fila, column=2, value="Saldo anterior").font = FUENTE_TOTALES
    celda = ws.cell(row=fila, column=5, value=estado.saldo_anterior)
    celda.number_format = "#,##0.00"
    celda.font = FUENTE_TOTALES
    fila += 1

    if not estado.movimientos:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
        ws.cell(
            row=fila,
            column=1,
            value="No se encontraron movimientos para el período seleccionado.",
        )
        fila += 1
    else:
        for m in estado.movimientos:
            c = m.comprobante
            celda_fecha = ws.cell(
                row=fila,
                column=1,
                value=datetime.strptime(c.fecha, "%Y-%m-%d").date(),
            )
            celda_fecha.number_format = "dd/mm/yyyy"
            ws.cell(
                row=fila,
                column=2,
                value=TIPOS_COMPROBANTE.get(c.tipo_cbte, str(c.tipo_cbte)),
            )
            ws.cell(row=fila, column=3, value=f"{c.punto_venta:04d}-{c.numero:08d}")
            celda_imp = ws.cell(row=fila, column=4, value=m.importe)
            celda_imp.number_format = "#,##0.00"
            celda_saldo = ws.cell(row=fila, column=5, value=m.saldo)
            celda_saldo.number_format = "#,##0.00"
            fila += 1

    ws.cell(row=fila, column=2, value=f"Saldo al {estado.fecha_fin}:").font = (
        FUENTE_TOTALES
    )
    celda_final = ws.cell(row=fila, column=5, value=estado.saldo_final)
    celda_final.number_format = "#,##0.00"
    celda_final.font = FUENTE_TOTALES

    anchos = {1: 12, 2: 20, 3: 16, 4: 16, 5: 16}
    for col in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = anchos.get(col, 16)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


ESTADO_LABEL = {"activos": "Activos", "inactivos": "Inactivos", "todos": "Todos"}


def generar_excel_listado_clientes(
    clientes: list[Cliente],
    *,
    texto: str,
    empresa: Empresa | None,
    estado: str,
) -> bytes:
    """Genera el XLSX del listado de clientes y devuelve los bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado de clientes"

    encabezado = [
        "Razón social",
        "Tipo Doc.",
        "Documento",
        "Cond. IVA",
        "Empresa",
        "Email",
        "Estado",
    ]
    n_cols = len(encabezado)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value="Listado de clientes").font = FUENTE_TITULO

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    subtitulo = f"Filtro: {ESTADO_LABEL.get(estado, estado)}"
    if texto:
        subtitulo += f"  |  Búsqueda: {texto}"
    ws.cell(row=2, column=1, value=subtitulo)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    ws.cell(
        row=3,
        column=1,
        value=f"Empresa: {empresa.razon_social if empresa else 'Todas las empresas'}",
    )

    fila_encabezado = 5
    for col, titulo in enumerate(encabezado, start=1):
        celda = ws.cell(row=fila_encabezado, column=col, value=titulo)
        celda.font = FUENTE_ENCABEZADO
        celda.fill = RELLENO_ENCABEZADO

    if not clientes:
        ws.merge_cells(
            start_row=fila_encabezado + 1,
            start_column=1,
            end_row=fila_encabezado + 1,
            end_column=n_cols,
        )
        ws.cell(
            row=fila_encabezado + 1,
            column=1,
            value="No se encontraron clientes para los filtros seleccionados.",
        )
    else:
        fila = fila_encabezado + 1
        for c in clientes:
            estado_texto = "Activo" if c.activo else f"Baja {c.fecha_baja:%d/%m/%Y}"
            ws.cell(row=fila, column=1, value=c.razon_social)
            ws.cell(row=fila, column=2, value=c.tipo_doc)
            ws.cell(row=fila, column=3, value=c.nro_doc)
            ws.cell(row=fila, column=4, value=c.condicion_iva)
            ws.cell(row=fila, column=5, value=c.empresa.razon_social)
            ws.cell(row=fila, column=6, value=c.email)
            ws.cell(row=fila, column=7, value=estado_texto)
            fila += 1

    anchos = {1: 35, 2: 10, 3: 16, 4: 22, 5: 28, 6: 30, 7: 18}
    for col in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = anchos.get(col, 16)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
